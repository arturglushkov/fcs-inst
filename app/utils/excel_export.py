"""Генерация Excel отчётов."""

import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


async def generate_excel_report(data: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    if not data:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    headers = list(data[0].keys())

    # Header row
    header_fill = PatternFill("solid", fgColor="1A5276")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)
    ws.row_dimensions[1].height = 28

    # Data rows
    for r, row in enumerate(data, 2):
        bg = "EBF5FB" if r % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=row.get(key, ""))
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                bottom=Side(style="thin", color="D5D8DC"),
                right=Side(style="thin", color="D5D8DC"),
            )
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
